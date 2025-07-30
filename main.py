import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

EXCEL_FILE = 'datos_registro.xlsx'

def guardar_datos():
    nombre = entry_nombre.get()
    edad = entry_edad.get()
    ciudad = entry_ciudad.get()

    if not nombre or not edad or not ciudad:
        messagebox.showerror("Error", "Todos los campos son obligatorios")
        return

    # Si el archivo existe, lo abre, si no, crea uno nuevo
    if os.path.exists(EXCEL_FILE):
        libro = load_workbook(EXCEL_FILE)
        ws = libro.active
    else:
        libro = Workbook()
        ws = libro.active
        ws.append(['Nombre', 'Edad', 'Ciudad'])

    ws.append([nombre, edad, ciudad])
    libro.save(EXCEL_FILE)
    messagebox.showinfo("Éxito", "Datos guardados correctamente")

root = tk.Tk()
root.title("Formulario de Registro")
root.config(bg="lightblue")

label_style = {
    "bg": "lightblue",
    "font": ("Arial", 12)
}

entry_style = {
    "bg": "white",
    "font": ("Arial", 12)
}

label_nombre = tk.Label(root, text="Nombre:", **label_style)
label_nombre.grid(row=0, column=0, padx=10, pady=10)
entry_nombre = tk.Entry(root, **entry_style)
entry_nombre.grid(row=0, column=1, padx=10, pady=10)

label_edad = tk.Label(root, text="Edad:", **label_style)
label_edad.grid(row=1, column=0, padx=10, pady=10)
entry_edad = tk.Entry(root, **entry_style)
entry_edad.grid(row=1, column=1, padx=10, pady=10)


label_ciudad = tk.Label(root, text="Ciudad:", **label_style)
label_ciudad.grid(row=2, column=0, padx=10, pady=10)
entry_ciudad = tk.Entry(root, **entry_style)
entry_ciudad.grid(row=2, column=1, padx=10, pady=10)

boton_guardar = tk.Button(root, text="Guardar Datos", command=guardar_datos, bg="green", fg="white", font=("Arial", 12))
boton_guardar.grid(row=3, column=0, columnspan=2, padx=10, pady=20)

root.mainloop()